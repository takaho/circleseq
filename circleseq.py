import os, sys, re, io, json, tempfile
import subprocess
import argparse
import numpy as np
import scipy.stats
import pandas as pd
import plotly
import plotly.graph_objs as go
import pymultiscale.anscombe
import statsmodels.stats.multitest
from logging import getLogger, Formatter, StreamHandler, FileHandler, DEBUG, INFO, ERROR, CRITICAL, FATAL

"""
CIRCLE-seq processing program.
CIRCLE-seq fastq files will be converted to statistically evaluated results using following commands.
Without specific commands, fastq files will be conevrted automatically. (map->count->compare->show)

map: convert fastq files into BAM files sorted by name
count : count reads obtained from genuine experiments.
compare : statitically evalauted with control experiments.
show : summarie counts and evaluated results into Excel and charts.
"""

def _convert_chrcode(chrom):
    chrom = re.sub('^chr', '', chrom)
    if chrom == 'X':
        return 50
    elif chrom == 'Y':
        return 51
    elif chrom.isdigit() is False:
        return 100
    else:
        return int(chrom)
def _reverse_complement(seq:str)->str:
    r = ''
    for i in range(len(seq)):
        r += {'A':'T', 'C':'G', 'G':'C', 'T':'A'}.get(seq[len(seq) - 1 - i], 'N')
    return r

def _load_tsv(filename)->dict:
    detection = {}
    with open(filename) as fi:
        fi.readline()
        for line in fi:
            items = line.strip().split('\t')
            chrom = items[4]
            start = int(items[5])
            stop = int(items[8])
            if re.match('chr[\\dXY]+$', chrom):
                if chrom not in detection: detection[chrom] = []
                detection[chrom].append((start + 1, 1))
    return detection

def _load_wig(filename)->dict:
    detection = {}
    chrom = None
    with open(filename) as fi:
        for line in fi:
            m = re.search('chrom=(\\w+)', line)
            if m:
                chrom = m.group(1)
                if re.match('chr[\\dXY]+$', chrom) is None:
                    chrom = None
            elif chrom is not None:
                items = line.strip().split('\t')
                if len(items) >= 2 and items[0].isdigit() and items[1].isdigit():
                    if chrom not in detection: detection[chrom] = []
                    detection[chrom].append((int(items[0]), int(items[1])))
    return detection

def _load_bed(filename)->dict:
    detection = {}
    with open(filename) as fi:
        for line in fi:
            if line.startswith('track'): continue
            items = line.strip().split('\t')
            if len(items) > 2:
                chrom = items[0]
                position = int(items[1])
                num = 1
            if len(items) > 3: # bedGraph
                num = int(items[3])
            if re.match('chr[\\dXY]+$', chrom):
                if chrom not in detection: detection[chrom] = []
                detection[chrom].append((position, num))
                # detection.append((chrom, position, num))
    return detection

def evaluate_enrichment(arguments=None, logger=None)->str:
    """Evaluate mapped reads
    Arguments:
        --track filenames (bed/wig/bedGraph)
        -w window size    
        -o output directory
        --verbose verbosity
        --without-control compare with background mappability
        --forced force calculation even if results are saved
    Return:
        statistics filename
    """
    if isinstance(arguments, argparse.Namespace):
        args = arguments
    else:
        parser = argparse.ArgumentParser()
        parser.add_argument('--track', metavar='filename', nargs='+', help='input filenames, control data should be placed first')
        parser.add_argument('-o', default='circleseq_out', metavar='directory', help='output directory')
        parser.add_argument('--verbose', action='store_true')
        parser.add_argument('--forced', action='store_true')
        parser.add_arguemnt('--without-control', action='store_true')
        parser.add_argument('--cutoff', type=int, default=10, help='cutoff counts of statistical evaluation')
        parser.add_argument('-w', type=int, default=1, metavar='number', help='window size', choices=range(1, 500))
        args = parser.parse_known_args(arguments)[0]
    logger = logger if logger is not None else getLogger() 
    outdir = args.o
    os.makedirs(outdir, exist_ok=True)

    titles = []
    detected = []
    allsites = {}
    window_size = args.w
    verbose = args.verbose
    if args.verbose:
        logger.setLevel(DEBUG)
    cutoff = args.cutoff
    wo_control = args.without_control
    forced = args.forced
    n_total = []
    fn_output = os.path.join(outdir, 'stat.tsv')
    fn_info = os.path.join(outdir, 'stat.info')
    with open(fn_info, 'w') as fo:
        fo.write('cutoff:{}\n'.format(cutoff))
        fo.write('window_size:{}\n'.format(window_size))
        fo.write('input_tracks:{}\n'.format(','.join(args.track)))

    track_filenames = args.track

    # load genome-wide observations
    n_inputs = len(track_filenames)
    n_test_sites = 0
    for file_index, filename in enumerate(track_filenames):
        ext = filename.split('.')[-1]
        logger.info('\033[Kloading "{}"'.format(filename))
        if ext == 'bed' or 'bedGraph':
            sites = _load_bed(filename)
        elif ext == 'wig':
            sites = _load_wig(filename)
        elif ext == 'tsv' or ext == 'csv':
            sites = _load_tsv(filename)
        else:
            sys.stderr.write('cannot load {}\n'.format(filename))
            continue
        title = os.path.basename(filename).split('.')[0]
        detected.append(sites)
        titles.append(title)
        n_reads = n_pos = 0
        for c in sites.keys():
            if c not in allsites:
                allsites[c] = {}
            for p, n in sites[c]:
                wn = (p // window_size) * window_size
                if wn not in allsites[c]:
                    allsites[c][wn] = np.zeros(n_inputs, dtype=np.int32)
                    n_test_sites += 1
                    n_pos += 1
                allsites[c][wn][file_index] += n
                n_reads += n
        n_total.append(n_reads)
        logger.info('{} loaded, title={}, n_pos={}, n_reads={}'.format(filename, title, n_pos, n_reads))

    logger.debug('{} cleavvage sites are under test'.format(n_test_sites))    

    if cutoff > 0:
        n_selected = 0
        for chrom in allsites.keys():
            filtered = {}
            for pos, cnt in allsites[chrom].items():
                if np.max(cnt) >= cutoff:
                    filtered[pos] = cnt
                    n_selected += 1
                    print(cnt)
            allsites[chrom] = filtered
        logger.debug('cutoff process selected {}/{} sites'.format(n_selected, n_test_sites))
        n_test_sites = n_selected

    chromosomes = list(sorted(allsites.keys()))
    n_inputs = len(detected)
    vst_total = pymultiscale.anscombe.anscombe(np.array(n_total, dtype=np.float32))
    count_table = []

    df_header = ['Chromosome', 'Start', 'Stop'] + titles
    if without_control:
        for i in range(n_inputs):
            df_header.append('lFC_{}'.format(titles[i]))
            df_header.append('pVal_{}'.format(titles[i]))
    else:
        for i in range(n_inputs):
            for j in range(i + 1, n_inputs):
                df_header.append('lFC_{}_{}'.format(titles[i], titles[j]))
                df_header.append('pVal_{}_{}'.format(titles[i], titles[j]))

    n_processed = 0
        
    for chromosome in chromosomes:
        for position in sorted(allsites[chromosome].keys()):
            n_mapped = allsites[chromosome][position]
            if cutoff > 0 and max(n_mapped) < cutoff:
                continue
            row = [chromosome, position, position + window_size] + list(n_mapped)

            try:
                tval = pymultiscale.anscombe.anscombe(n_mapped)#np.array(n_mapped))
            except Exception as e:
                sys.stderr.write(repr(n_mapped) + ' ' + repr(e) + '\n')
                continue
            if without_control:
                ####################################################
                for i in range(n_inputs):
                    total = vst_total[i]
                    if n_mapped[i] > 0: #
                        logfc = np.log2(n_counts[i] * n_sites[i] / n_total[i])
                        pvalue = scipy.stats.nbinom.cdf(vst_total[i], tval[i], tval[i] / vst_total[i])
                        pass
                    else:
                        logfc = 0.0
                        pvalue = 1.0
                    row += [logfc, pvalue]
            else: 
                for i in range(n_inputs):
                    for j in range(i + 1, n_inputs):
                        logfc = np.log2(tval[j] * vst_total[i] / (tval[i] * vst_total[j]))
                        if logfc > 0:
                            prob_ = tval[i] / vst_total[i]
                            obs_ = tval[j]
                            # print('prob, obs={:.3f}, {:.3f}'.format(prob_, obs_))
                            pvalue = scipy.stats.nbinom.cdf(vst_total[j], obs_, prob_)
                        else:
                            prob_ = tval[j] / vst_total[j]
                            obs_ = tval[i]
                            pvalue = scipy.stats.nbinom.cdf(vst_total[i], obs_, prob_)
                        # if pvalue < 1e-200:
                        #     logpvalue = 200
                        # else:
                        #     logpvalue = -np.log10(pvalue)
                        # print(i, j, logfc, pvalue, tval[j], tval[i])
                        row += [logfc, pvalue]

            count_table.append(row)
            if logger.getEffectiveLevel() <= DEBUG:
                sys.stderr.write('\033[K {:.1f}% {}:{}\r'.format(n_processed * 100 / n_test_sites, chromosome, position))
            n_processed += 1
    if logger.getEffectiveLevel() <= DEBUG:
        sys.stderr.write('\033[K\r')    

    df = pd.DataFrame(count_table, columns=df_header)
    dfs = []
    for i, c in enumerate(df.columns):
        dfs.append(df[c])
        if c.startswith('pVal_'):
            pvalues = df[c].values.reshape(-1)
            mres = statsmodels.stats.multitest.multipletests(pvalues)
            qvalue = pd.DataFrame(mres[1].reshape(-1, 1), index=df.index, columns=['qVal_' + c[5:]])
            dfs.append(qvalue)
    df = pd.concat(dfs, axis=1)
    df.to_csv(fn_output, sep='\t')
    return fn_output

def load_blacklist(filename:str, **kwargs)->dict:
    """Load UCSC blacklist of gzipped bed file.
    Return : dictionary (key=chromosome, value=list of (start, stop))
    """
    import gzip, io
    verbose = kwargs.get('verbose', False)
    logger = kwargs.get('logger', getLogger())
    if verbose: logger.setLevel(DEBUG)
    if filename.endswith('.gz'):
        istr = io.TextIOWrapper(gzip.open(filename), encoding='utf-8')
    else:
        istr = open(filename)
    bl = {}
    chromosomes = {}
    n_bases = 0
    n_regions = 0
    for line in istr:
        items = line.strip().split('\t')
        if len(items) >= 3 and items[0].startswith('chr'):
            start = int(items[1])
            stop = int(items[2])
            group = items[3]
            if group not in bl:bl[group] = 0
            bl[group] += stop - start
            chrom = items[0]
            n_bases += stop - start
            n_regions += 1
            if chrom not in chromosomes:
                chromosomes[chrom] = [(start, stop)]
            else:
                chromosomes[chrom].append((start, stop))
    # if verbose:
    n_chrom = len(chromosomes)
    logger.info('Blacklist on {} chromosomes, {} sites, {} bases'.format(n_chrom, n_regions, n_bases))
        # sys.stderr.write('Blacklist\tsize\n')
        # for group in bl.keys():
        #     sys.stderr.write('{}\t{}\n'.format(group, bl[group]))
    istr.close()
    return chromosomes

def load_chromosome_sizes(build_version:str='mm10', *, cache:str=None, logger=None)->dict:
    """Load chromosome size of certain build version from UCSC
    Return : dictionay {'size':dictionary of chromosome and length pair, 
    'filename':'container filename'}
    """
    logger = logger if logger is not None else getLogger()
    if cache is None:
        cache_dir = os.path.expanduser('~/.chrom.sizes')
        os.makedirs(cache_dir, exist_ok=True)
        cache = os.path.join(cache_dir, build_version + '.chrom.sizes')
    chrom = {}
        
    if os.path.exists(cache):
        with open(cache) as fi:
            for line in fi:
                items = line.strip().split('\t')
                if len(items) >= 2 and items[1].isdigit():
                    chrom[items[0]] = int(items[1])
    if len(chrom) > 5 and sum(chrom.values()) > 1000:
        return {'filename':cache,  'size':chrom}
    import urllib3
    http = urllib3.PoolManager()
    url = 'http://hgdownload.soe.ucsc.edu/goldenPath/{0}/bigZips/{0}.chrom.sizes'.format(build_version)
    res = http.request('GET', url)
    if res.status == 200:
        contents = res.data
        with open(cache, 'wb') as fo:
            fo.write(contents)
        with open(cache) as fi:
            for line in fi:
                items = line.strip().split('\t')
                if len(items) >= 2 and items[1].isdigit():
                    chrom[items[0]] = int(items[1])
        return {'filename':cache,  'size':chrom}
    else:
        logger.error('cannot load chromosome size from {}'.format(url))
        raise Exception('cannot load contents from {}'.format(url))

def count_cutsites(arguments=None, *, logger=None):
    """Read SAM/BAM file and count mapped tags, alignment files *MUST BE SORTED BY NAME* """
    logger = getLogger() if logger is None else logger
    if isinstance(arguments, argparse.Namespace):
        args = arguments
    else:
        parser = argparse.ArgumentParser()
        parser.add_argument('-b', default=[], metavar='filename...', nargs='+', help='BAM or SAM files, sorted by name (samtoosl sort -n)' )
        parser.add_argument('-o', default='out', metavar='direcotry', help='output directory')
        parser.add_argument('--tolerance', type=int, default=5, help='distance tolerance (read size * 2 +- tolerance)')
        parser.add_argument('--cutoff', default=0, type=int, help='cutoff threshold')
        parser.add_argument('--forced', action='store_true')
        parser.add_argument('--verbose', action='store_true')
        parser.add_argument('--blacklist', metavar='filename', help='blacklist')
        args = parser.parse_known_args(arguments)[0]
    outdir = args.o
    verbose = args.verbose
    tolerance = args.tolerance
    cutoff = args.cutoff
    forced = args.forced
    fn_blacklist = args.blacklist
    os.makedirs(outdir, exist_ok=1)
    fn_info = os.path.join(outdir, 'count.info')
    bamfiles = args.b

    if verbose:
        logger.setLevel(DEBUG)

    output_filenames = []

    if fn_blacklist is not None and os.path.exists(fn_blacklist):
        blacklist = load_blacklist(fn_blacklist, verbose=verbose, logger=logger)
    else:
        blacklist = {}

    with open(fn_info, 'w') as fo:
        fo.write('inputs:{}\n'.format(','.join(args.b)))
        fo.write('cutoff:{}\n'.format(cutoff))
        fo.write('overlap_tolerance:{}\n'.format(tolerance))
        if fn_blacklist:
            fo.write('blacklist:{}\n'.format(fn_blacklist))

    traces = []
    xpos = np.arange(0, 101, 1)

    # Load BAM files and convert to bedgraph
    file_index = 0
    for fn in bamfiles:
        label = os.path.basename(fn).split('.')[0]
        fn_out = os.path.join(outdir, label + '.bedgraph')
        fn_dist = os.path.join(outdir, label + '.tsv')
        logger.info('{}/{} : {}'.format(file_index + 1, len(bamfiles), fn))

        if not forced and os.path.exists(fn_out) and os.path.getsize(fn_out) > 100:
            if verbose:
                sys.stderr.write('skip {}\n'.format(fn))
            continue

        if fn.endswith('.bam'):
            cmd = 'samtools', 'view', '-h', fn
            proc = subprocess.Popen(cmd, stdout=subprocess.PIPE)
            fi = io.TextIOWrapper(proc.stdout, encoding='ascii')
        else:
            proc = None
            fi = open(fn)
        sequences = {}
        current_seqid = ''
        current_chrom = ''
        maps = []

        n_reads = 0
        n_pairs = 0
        n_mapped = 0
        n_sites = 0
        results = {}
        last_pos = -1
        for line in fi:
            if line.startswith('@'):
                m = re.search('@SQ\\s+SN:(\\S+)\\s+LN:(\\d+)', line)
                if m:
                    sequences[m.group(1)] = int(m.group(2))
                    results[m.group(1)] = []
            else:
                # read mapped
                n_reads += 1
                items = line.split('\t')
                seqid = items[0]
                chrom = items[2]
                pair_chrom = items[6]
                if seqid != current_seqid: # new tag
                    if len(maps) > 0:# and len(maps) <= 2:
                        cleavage_site = detect_cleavage_site(maps, current_chrom, tolerance=tolerance, verbose=verbose)
                        if cleavage_site >= 0:
                            n_pairs += 1
                            results[current_chrom].append(cleavage_site)
                        maps = [] 
                    current_seqid = seqid
                    current_chrom = chrom
                if chrom != '*' and pair_chrom == '=':
                    n_mapped += 1
                    flag = int(items[1])
                    # chrom = items[2]
                    pos = int(items[3])
                    distance = int(items[8])
                    cigar = items[5]
                    m = re.match('(\\d+)([A-Z])', cigar)
                    max_span = 0
                    while m:
                        op = m.group(2)
                        span = int(m.group(1))
                        if op == 'M':
                            max_span = max(span, max_span)
                        cigar = cigar[m.end():]
                        m = re.match('(\\d+)([A-Z])', cigar)
                    size = max_span
                    maps.append((chrom, pos, flag, distance, size))
            if verbose and n_reads % 1000 == 0:
                sys.stderr.write('\033[K{}\t{}\t{}k\t{}\r'.format(current_seqid, label, n_reads // 1000, n_pairs))
        if len(maps) > 0:
            cleavage_site = detect_cleavage_site(maps, current_chrom, tolerance=tolerance, verbose=verbose)
            if cleavage_site >= 0:
                results[current_chrom].append(cleavage_site)
        
        # remove counts in blacklist
        if len(blacklist) > 0:
            logger.info('removing blacklist')
        n_removed_total = 0
        for chrom in results.keys():
            if chrom not in blacklist: continue
            positions = sorted(results[chrom])
            flags = [False] * len(positions)
            index_start = 0
            index_stop = len(positions)
            i = 0
            for start, stop in sorted(blacklist[chrom], key=lambda p:p[0]):
                left = 0
                right = len(positions)
                pivot = -1
                while left < right:
                    center = (left + right) // 2
                    p = positions[center]
                    if p < start:
                        left = center + 1
                    elif p > stop:
                        right = center
                    else:
                        pivot = center
                        break
                if pivot >= 0:
                    i = pivot
                    while i >= 0:
                        p = positions[i]
                        if start <= p <= stop:
                            flags[i] = True
                        else:
                            break
                        i -= 1
                    i = pivot + 1
                    while i < len(positions):
                        p = positions[i]
                        if start <= p <= stop:
                            flags[i] = True
                        else:
                            break
                        i += 1
            n_removed_total += sum(flags)
            results[chrom] = [positions[i] for i in range(len(positions)) if not flags[i]]
        if n_removed_total > 0:
            logger.info('{} reads in blacklist regions were removed'.format(n_removed_total))
        freq = {}
        output_filenames.append(fn_out)
        n_sites = n_sites_plus = 0
        logger.info('writing bedGraph')
        with open(fn_out, 'w') as fo:
            name = '{};total={};tolerance={}'.format(label, n_pairs, tolerance)
            if cutoff > 0:
                name += ';cutoff={}'.format(cutoff)
            fo.write('track type=bedGraph name="{}"\n'.format(name))
            for chrom in sorted(results):
                loc = results[chrom]
                cnt = {}
                for pos in loc:
                    cnt[pos] = cnt.get(pos, 0) + 1
                n_sites += sum(cnt.values())
                for p in sorted(cnt.keys()):
                    n = cnt[p]
                    freq[n] = freq.get(n, 0) + 1
                    if n > cutoff:
                        n_sites_plus += n
                        fo.write('{}\t{}\t{}\t{}\n'.format(chrom, p, p + 1, n))
        logger.info('writing statistics')
        with open(fn_dist, 'w') as fo:
            fo.write('num_reads\tcount\n')
            for c in sorted(freq.keys()):
                fo.write('{}\t{}\n'.format(c, freq[c]))
        fi.close()
        logger.info('{} contains {} reads, {} pairs, {} sites'.format(fn, n_reads, n_pairs, n_sites))
        if proc is not None:
            proc.stdout.close()
            proc.wait()

        with open(fn_info, 'a') as fo:
            fo.write('n_reads:{}:{}\n'.format(label, n_reads))            
            fo.write('n_cleavage_pairs:{}:{}\n'.format(label, n_pairs))
            if cutoff > 0:
                fo.write('n_cleavage_sites_above:{}:{}\n'.format(label, n_sites_plus))
            fo.write('n_cleavage_sites:{}:{}\n'.format(label, n_sites))
    return output_filenames # returns bedgraph filenames

def detect_cleavage_site(reads, chromosome:str, tolerance:int=4, max_distance:int=2000, *, verbose:bool=False, logger=None)->int:
    """Detection of CIRCLE-seq cleavage sites using positions of mapped reads and 
    SAM file flag information
    If multiple sites were proposed by the alignments, most frequenct location is proposed.

    Return : Detected position
    """
    logger = logger if logger is not None else getLogger()
    if verbose:
        logger.setLevel(DEBUG)
    # divide into first/second pair group
    if len(reads) < 2:
        return -1
    first = []
    second = []
    for chrom, pos, flag, distance, size in sorted(reads, key=lambda x_:x_[1]):
        if chrom != chromosome: continue
        reverse_strand = (flag & 16) != 0
        first_in_pair = (flag & 64) != 0
        mate_reverse = (flag & 32) != 0
        if mate_reverse != reverse_strand:
            if first_in_pair: # first in pair
                first.append((chrom, pos, flag, distance, size))
            else:
                second.append((chrom, pos, flag, distance, size))
    positions = {}
    for chrom, pos, flag, distance, size in first:
        for c_, p_, f_, d_, s_ in second:
            if distance == -d_ and chrom == c_ and abs(pos - p_) < max_distance:
                if pos < p_: # pos <-> c <-> p_
                    cleavage = pos + size
                    if cleavage - tolerance > p_:
                        continue
                else: # p_ <-> c <-> pos
                    cleavage = p_ + s_
                    if cleavage - tolerance > pos:
                        continue
                positions[cleavage] = positions.get(cleavage, 0) + 1#.append(cleavage)
    if len(positions) == 0: return -1
    maxcnt = max(positions.values())
    sites = [p for p in positions.keys() if positions[p] == maxcnt]
    return sites[np.random.randint(0, len(sites))]

def load_chromosome_size(build_version='hg38', *, cache=None, logger=None):
    """Load chromosome size of certain build version from UCSC"""
    logger = logger if logger is not None else getLogger()
    if cache is None:
        cache_dir = os.path.expanduser('~/.chrom.sizes')
        os.makedirs(cache_dir, exist_ok=True)
        cache = os.path.join(cache_dir, build_version + '.chrom.sizes')
    chrom = {}
        
    if os.path.exists(cache):
        with open(cache) as fi:
            for line in fi:
                items = line.strip().split('\t')
                if len(items) >= 2 and items[1].isdigit():
                    chrom[items[0]] = int(items[1])
    if len(chrom) > 5 and sum(chrom.values()) > 1000:
        return chrom
    import urllib3
    http = urllib3.PoolManager()
    url = 'http://hgdownload.soe.ucsc.edu/goldenPath/{0}/bigZips/{0}.chrom.sizes'.format(build_version)
    res = http.request('GET', url)
    if res.status == 200:
        contents = res.data
        with open(cache, 'wb') as fo:
            fo.write(contents)
        with open(cache) as fi:
            for line in fi:
                items = line.strip().split('\t')
                if len(items) >= 2 and items[1].isdigit():
                    chrom[items[0]] = int(items[1])
        return chrom
    else:
        raise Exception('cannot load contents from {}'.format(url))

def display_results(arguments=None, *, logger=None)->dict:    
    """Draw significantly enriched positions.
    Return : {excel:excel filename, 
              chart:PDF filename}
    """
    logger = logger if logger is not None else getLogger()
    import reportlab.pdfgen.canvas
    import openpyxl
    import subsequence

    if isinstance(arguments, argparse.Namespace):
        args = arguments
    else:
        parser = argparse.ArgumentParser()
        parser.add_argument('--verbose', action='store_true')
        parser.add_argument('-i',help='output of compare command')
        parser.add_argument('-g', default='hg38', help='genome build (default hg38)')
        parser.add_argument('-o', default=None, help='output directory', metavar='directory')
        parser.add_argument('--cutoff', type=int, default=10)    
        parser.add_argument('--column', nargs='+', type=int, default=[])
        parser.add_argument('-q', type=float, default=0.01)
        parser.add_argument('-l', type=float, default=0.5)
        parser.add_argument('--geometry', default='800x300')
        parser.add_argument('--span', type=int, default=16)
        parser.add_argument('--title', default=None)
        parser.add_argument('--fasta', default=None, metavar='filename', help='fasta format genome')
        args = parser.parse_known_args(arguments)[0]

    qthr = args.q    
    lthr = args.l
    verbose = args.verbose
    filename_tsv = args.i    
    filename_fasta = args.fasta
    title = args.title
    geom = [int(x_) for x_ in args.geometry.split('x')[0:2]]
    width = max(0, min(1000, geom[0]))
    height = max(0, min(1000, geom[1]))
    margin = min(100, max(25, width / 4))
    span = args.span
    if args.verbose:
        logger.setLevel(DEBUG)
    if title is None: title = re.split('\\W', os.path.basename(filename_tsv))[0]

    outdir = args.o
    if outdir is None:
        outdir = os.path.dirname(filename_tsv)
    else:
        os.makedirs(outdir, exist_ok=True)

    filename_info = os.path.join(outdir, 'display.info')
    filename_chart = os.path.join(outdir, '{}_cleavagesites.pdf'.format(title))
    filename_excel = os.path.join(outdir, '{}_cleavagesites.xlsx'.format(title))
    genome = args.g

    chrlength = {}
    offset = {}
    cutoff = args.cutoff
    chromosome_size = load_chromosome_size(genome)

    with open(filename_info, 'w') as fo:
        fo.write('input:{}\n'.format(filename_tsv))
        fo.write('genome:{}\n'.format(genome))
        if filename_fasta:
            fo.write('genome_seq:{}\n'.format(args.fasta))
        fo.write('width:{}\n'.format(width))
        fo.write('height:{}\n'.format(height))

    # define chromosome sizes and offset positions
    chromosome_labels = list(sorted(chromosome_size.keys(), key=lambda c_:_convert_chrcode(c_)))
    ofval = 0
    for ch in chromosome_labels:
        if re.match('chr[\\dXY]+$', ch):
            chrlength[ch] = chromosome_size[ch]
            offset[ch] = ofval
            ofval += chromosome_size[ch]
    total_length = ofval

    # add genome
    df_results = pd.read_csv(filename_tsv, sep='\t', index_col=0)

    lfc_cols = []
    qval_cols = []
    pval_cols = []
    for i, col in enumerate(df_results.columns):
        if col.startswith('qVal_'):
            qval_cols.append(i)
        elif col.startswith('pVal_'):
            pval_cols.append(i)
        elif col.startswith('lFC_'):
            lfc_cols.append(i)
    n_tests = min(len(lfc_cols), len(qval_cols), len(pval_cols))
    n_samples = min(lfc_cols + qval_cols + pval_cols) - 1 - 2
    sample_names = df_results.columns[3:3+n_samples]

    significant = []
    results = []
    book = openpyxl.workbook.Workbook()
    book.remove(book.active)
    sheet = book.create_sheet('Detected')
    sheet.append(list(df_results.columns) + ['Up', 'Down', 'Down_cmp', 'Up_cmp'])
    detection_marks = [[[], []] for i in range(n_tests)]
    redfont = openpyxl.styles.Font(color='00FF0000')

    row_index = 2
    for i, row in enumerate(df_results.values):
        flag = False
        significant = []
        for j in range(n_tests):
            lfc = row[lfc_cols[j]]
            qval = row[qval_cols[j]]
            # print(lfc)
            # print(qval)
            if qval < qthr:
                if abs(lfc) > lthr:
                    significant.append(j)
                    detection_marks[j][lthr < 0].append(i) # mark
        if len(significant) > 0:
            sheet.cell(row=row_index, column=2).number_format = '0'
            sheet.cell(row=row_index, column=3).number_format = '0'
            for col, val in enumerate(row):
                sheet.cell(row=row_index, column=col + 1).value = val
            for col in qval_cols + pval_cols:
                sheet.cell(row=row_index, column=col + 1).number_format = '0.00E+00'
            for col in lfc_cols:
                sheet.cell(row=row_index, column=col + 1).number_format = '0.00'
            for j in significant:
                sheet.cell(row=row_index, column=lfc_cols[j] + 1).font = redfont
                sheet.cell(row=row_index, column=pval_cols[j] + 1).font = redfont
                sheet.cell(row=row_index, column=qval_cols[j] + 1).font = redfont

            if filename_fasta:
                chrom, start, stop = row[0:3]
                pos = (start + stop) // 2
                nucleotide = subsequence.get_subsequence(filename_fasta, chrom, pos - span, pos + span)
                if len(nucleotide) < span * 2:
                    nucleotide += ' ' * (span * 2 - len(nucleotide))
                seqs = [nucleotide[0:span], 
                    nucleotide[span:]]
                seqs += [_reverse_complement(seqs[1]),
                    _reverse_complement(seqs[0])]
                for j in range(4):
                    sheet.cell(row=row_index, column=j + len(row) + 1).value = seqs[j]
            row_index += 1
    book.save(filename_excel)

    # Draw charts
    index = 0

    pagesize = (margin + (margin + width) * 2, margin + (height + margin) * n_tests)
    verbose = True
    if verbose:
        sys.stderr.write('chart {}x{}x{}: {}\n'.format(width, height, n_tests, filename_chart))
    cnv = reportlab.pdfgen.canvas.Canvas(filename_chart, pagesize=pagesize)
    left, right = margin, margin + width
    index_chart = 0
    chromosome2color = {}
    barcolors = ['ReportLabFidBlue', 'ReportLabFidRed', 'ReportLabGreen', 'purple', 'orange', 
            'gold', 'darkslateblue', 'brown', 'cyan', 'magenta', 'olive', 'gray', 'pink']
    for chrom in sorted(offset.keys(), key=lambda c:offset[c]):
        chromosome2color[chrom] = barcolors[len(chromosome2color) % len(barcolors)]
        pass
    location = df_results.iloc[:,0:2].values.reshape(-1, 2)
    for index_0, index_1 in [(x_,y_) for x_ in range(n_samples) for y_ in range(n_samples)]:
        if index_0 >= index_1: continue
        count_0 = df_results[df_results.columns[3+index_0]].values.reshape(-1)
        count_1 = df_results[df_results.columns[3+index_1]].values.reshape(-1)
        qval = df_results[df_results.columns[qval_cols[index_chart]]].values.reshape(-1)
        lfc = df_results[df_results.columns[lfc_cols[index_chart]]].values.reshape(-1)
        # print(location.shape)
        for i in range(2):
            cnv.saveState()
            count = [count_0, count_1][1-i]
            ymax = np.max(count_0) if i == 1 else np.max(count_1)
            fig = np.log10(ymax + 1)
            c = np.floor(fig)
            r = fig - c
            if r < np.log10(2):
                ymax = 2 * (10 ** c)
            elif r < np.log10(5):
                ymax = 5 * (10 ** c)
            else:
                ymax = 10 ** (c + 1)
            xcnv = lambda chrom, pos: (offset[chrom] + pos) / total_length * width + margin + (i * (margin + width))
            ycnv = lambda cnt: cnt / ymax * height + pagesize[1] - (index_chart + 1) * (height + margin)
            x0 = xcnv('chr1', 0)
            x1 = x0 + width
            y0 = ycnv(0)
            y1 = ycnv(ymax)
            used = set()
            for j in range(df_results.shape[0]):
                if qval[j] >= qthr: continue
                if lfc[j] * (1 if i == 0 else -1) > lthr:
                    chromosome, pos = location[j]
                    if chromosome not in offset: continue
                    x = xcnv(chromosome, pos)
                    color = chromosome2color[chromosome]
                    cnv.setFillColor(color)
                    y = ycnv(count[j])
                    cnv.rect(x - .5, y0, 1, y - y0, 0, 1)
                    used.add(chromosome)
                    logger.info('{}\t{}\t{}:{}\t{}\t{}'.format(i, j, chromosome, pos, count[j], y - y0))
            cnv.setStrokeColor('black')
            cnv.setFillColor('black')
            graph_label = '{} {} {}'.format(sample_names[index_0], '<>'[i], sample_names[index_1])
            cnv.drawString(x0, y1 + 10, graph_label)
            ymax_ = '{:.0f}'.format(ymax)
            cnv.drawString(x0 - 10 - cnv.stringWidth('0'), y0, '0')
            cnv.drawString(x0 - 10 - cnv.stringWidth(str(ymax_)), y1, ymax_)
            for c in offset.keys():
                if c not in used: continue
                cnv.setFillColor(chromosome2color[c])
                cnv.setFont('Helvetica', 8)
                x_l = xcnv(c, 0)
                x_r = xcnv(c, chromosome_size[c])
                cnv.saveState()
                cnv.translate((x_l + x_r) / 2, y0)
                cnv.rotate(-90)
                cnv.drawString(10, 0, c)
                # cnv.drawString((x_l + x_r - cnv.stringWidth(c)) / 2, y0 - 10, c)
                cnv.restoreState()
            cnv.setLineWidth(1)
            cnv.rect(x0, y0, x1 - x0, y1 - y0)
            cnv.restoreState()
        index_chart += 1    
        pass
    cnv.save()
    return {'chart':filename_chart, 'spreadsheet':filename_excel}

def _get_code(params):
    import hashlib
    md5 = hashlib.md5()
    for key in sorted(params.keys()):
        val = params[key]
        md5.update('{}={};'.format(str(key), str(val)).encode('utf8'))
    code = md5.hexdigest()
    return code

def align_reads(arguments=None, logger=None):
    """Align paired-end reads on genome and return BAM file(s) sorted by sequence ID.
    """
    logger = logger if logger is not None else getLogger()
    if isinstance(arguments, argparse.Namespace):
        args = arguments
    else:
        parser = argparse.ArgumentParser()
        parser.add_argument('--r1', nargs='+', metavar='R1 fastq file', help='R1 sequence')
        parser.add_argument('--r2', nargs='+', metavar='R2 fastq file', help='R2 sequence')
        parser.add_argument('--fastq', nargs='+', metavar='fastq files', help='automatic detection of R1/R2 sequences')
        parser.add_argument('-o', metavar='directory', default='circleseq_out')
        parser.add_argument('--bowtie2-path', default='bowtie2')
        parser.add_argument('--bowtie2-db', help='bowtie2 database', metavar='hg38.bt2')
        parser.add_argument('--verbose', action='store_true')
        parser.add_argument('--title', default=None, metavar='string', help='title of the data')
        parser.add_argument('--forced', action='store_true', help='force alignment even if BAM file exists')
        parser.add_argument('-p', type=int, default=0)
        code = None
        args = parser.parse_args(arguments)[0]

    verbose = args.verbose
    title = args.title
    
    bt2 = args.bowtie2_path
    bt2_db = args.bowtie2_db
    forced = args.forced
    outdir = args.o
    num_threads = args.p
    if num_threads is None or num_threads <= 0:
        import multiprocessing
        num_threads = multiprocessing.cpu_count()

    # automatic arrange R1 and R2 sequences
    r1seq = args.r1
    r2seq = args.r2
    titles = []
    if r1seq is None or r2seq is None:
        if args.fastq is None:
            logger.error('no fastq files given')
            raise Exception('no fastq files')
        import collections
        r1seq = []
        r2seq = []
        fastqfiles = collections.OrderedDict()
        for fn in args.fastq:
            m = re.search('(.*)_R(1|2)\\D.*\\.fastq(\\.gz)?$', os.path.basename(fn))
            if m:
                label = m.group(1)
                rnum = m.group(2)
                if label not in fastqfiles: fastqfiles[label] = [None, None]
                fastqfiles[label][int(rnum) - 1] = fn
        for label in fastqfiles.keys():
            r1_, r2_ = fastqfiles[label]
            if r1_ is not None and r2_ is not None and os.path.exists(r1_) and os.path.exists(r2_):
                titles.append(label)
                r1seq.append(r1_)
                r2seq.append(r2_)
                logger.info('{} <- {}/{}'.format(label, os.path.basename(r1_), os.path.basename(r2_)))
    else:
        r1seq = args.r1
        r2seq = args.r2
        titles = [os.path.basename(f).split('_')[0] for f in args.r1]

    bamfiles = []
    for index in range(len(r1seq)):
        r1 = r1seq[index]
        r2 = r2seq[index]
        bam_title = titles[index]
        if bt2_db is None:
            logger.error("no bowtie2 database is given")
            raise Exception('no bt2 database')

        os.makedirs(outdir, exist_ok=True)
        samfile = os.path.join(outdir, bam_title + '.sam')
        bamfile = os.path.join(outdir, bam_title + '.bam')

        if forced or (os.path.exists(samfile) is False or os.path.getsize(samfile) < 1000):
            cmd = (bt2, '-x', bt2_db, '-p', str(num_threads),
                '-1', r1, '-2', r2, '-S', samfile, '--very-sensitive-local')
            logger.info(' '.join(cmd))
            proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
            for line in proc.stdout:
                logger.info(line)
            errocode = proc.wait()
            proc.stdout.close()
            del proc
            if errorcode < 0:
                logger.error("bowtie2 eror with code {}".format(errorcode))
                os.unlink(samfile)
            if os.path.exists(samfile) is False:
                raise Exception('failure of alignment\n')

        if forced or (os.path.exists(bamfile) is False or os.path.getsize(bamfile) < 1000):
            cmd = 'samtools', 'sort', '-n', '-o', bamfile, samfile
            logger.info(' '.join(cmd))
            err = subprocess.Popen(cmd).wait()
            if err < 0:
                logger.error("samtools sort faled with code {}".format(err))
                os.unlink(bamfile)
            if os.path.exists(bamfile) is False:
                raise Exception('failure of conversion to bam')
        bamfiles.append(bamfile)

    return bamfiles

def main():
    """
    command map/count/eval/show
    """

    logger = getLogger(__name__)
    parser = argparse.ArgumentParser()

    preferences = {
        'cutoff':10,
        'threshold_qvalue':0.01,
        'threshold_logfc':0.5,
        'genome':'hg38',
        'output':'circleseq_out',
        'bowtie2':'bowtie2',
        'bowtie2-db':'hg38',
    }

    for dirname in [os.path.dirname(__file__), '~', '.']:
        fn = os.path.expanduser(os.path.join(dirname, '.circleseq.json'))
        if os.path.exists(fn) and os.path.getsize(fn) is False:
            info = json.load(open(fn))
            if isinstance(info, dict):
                for key, val in info.items():
                    preferences[key] = val

    # common
    parser.add_argument('--verbose', action='store_true')
    parser.add_argument('--forced', action='store_true', help='force alignment even if BAM file exists')
    parser.add_argument('-g', default=preferences['genome'], help='genome build (default hg38)')
    parser.add_argument('-o', default=preferences['output'], help='output directory', metavar='directory')
    parser.add_argument('-p', type=int, default=0, metavar='CPU', help='number of processors')

    # Alignment
    parser.add_argument('--r1', nargs='+', metavar='R1 fastq files', help='Alignment: R1 sequence')
    parser.add_argument('--r2', nargs='+', metavar='R2 fastq files', help='Alignment: R2 sequence')
    parser.add_argument('--fastq', nargs='+', metavar='fastq files', help='automatic detection of R1/R2 sequences')
    parser.add_argument('--bowtie2-path', default='bowtie2')
    parser.add_argument('--bowtie2-db', help='bowtie2 database', metavar='hg38.bt2')
    parser.add_argument('--title', default=None, metavar='string', help='title of the data')

    # count
    parser.add_argument('-b', default=[], metavar='filename...', nargs='+', help='SAM/BAM files, sorted by name (samtoosl sort -n)' )
    parser.add_argument('--tolerance', type=int, default=5, help='distance tolerance (read size * 2 +- tolerance)')
    parser.add_argument('--cutoff', default=0, type=int, help='cutoff threshold')
    parser.add_argument('--blacklist', metavar='filename', help='blacklist')

    # evaluate
    parser.add_argument('--without-control', help='evaluate count without control')
    parser.add_argument('--track', nargs='+', metavar='bed/bedGraph/wig filenames', help='whole-genome track of cleavage sites')
    parser.add_argument('-w', type=int, default=5, metavar='int', help='window size of analysis, >0')

    # show display
    parser.add_argument('-i', metavar='filename', help='output of compare command')
    parser.add_argument('-q', type=float, default=0.01)
    parser.add_argument('-l', type=float, default=0.5)
    parser.add_argument('--fasta', default=None, metavar='filename', help='fasta format genome')
    parser.add_argument('--geometry', default='800x300')
    parser.add_argument('--span', type=int, default=16)

    parser.add_argument('--log', default=None)

    args, cmds = parser.parse_known_args()
    forced = args.forced
    outdir = args.o
    os.makedirs(outdir, exist_ok=True)

    fn_info = os.path.join(outdir, 'run.info')
    if os.path.exists(fn_info):
        with open(fn_info) as fi:
            for key, val in json.load(fi).items():
                preferences[key] = val
    else:
        preferences = vars(args)

    processed = False
    verbose = args.verbose

    # set logging
    def _set_log_handler(logger, handler):#, verbose):
        handler.setFormatter(Formatter('%(asctime)s %(name)s:%(lineno)s %(funcName)s [%(levelname)s]: %(message)s'))
        logger.addHandler(handler)
        return logger
    logfile= args.log
    if logfile is None:
        import tempfile
        logfile = tempfile.mktemp()#dtemp()
    _set_log_handler(logger, StreamHandler())
    _set_log_handler(logger, FileHandler(logfile))
    if verbose:
        logger.setLevel(DEBUG)
    else:
        logger.setLevel(ERROR)
    logger.propagate = False

    try:
        if 'map' in cmds: 
            bamfiles = align_reads(args, logger=logger)
            if 'b' in args:
                sys.stderr.write('BAM files were overwritten as generated.')
            args.b = bamfiles
            processed = True
        if 'count' in cmds:
            bedgraph_files = count_cutsites(args, logger=logger)
            args.track = bedgraph_files        
            processed = True
        if 'eval' in cmds:
            summary_file = evaluate_enrichment(args, logger=logger)
            args.i = summary_file
            processed = True
        if 'show' in cmds:
            display_results(args, logger=logger)
            processed = True
        if not processed: # do all
            if 'b' not in args or args.b is None or len(args.b) == 0 or os.path.exists(args.b[0]) is False or forced:
                sys.stderr.write('aligning...\n')
                bamfiles = align_reads(args, logger=logger)
                args.b = bamfiles
            if 'track' not in args or args.track is None or len(args.track) == 0 or os.path.exists(args.track[0]) is False or forced:
                sys.stderr.write('track...\n')
                bedgraph_files = count_cutsites(args, logger=logger)
                args.track = bedgraph_files        
            if 'i' not in args or args.i is None or forced:
                sys.stderr.write('-i argment was ignored')
                summary_file = evaluate_enrichment(args, logger=logger)
                args.i = summary_file
            display_results(args)
        with open(fn_info, 'w') as fo:
            json.dump(vars(args), fo, indent=2)
    except Exception as e:
        logger.error('Process interrupted : {}'.format(str(e)))
        raise
if __name__ == '__main__':
    main()
